import os
import time
import asyncio
import concurrent.futures
import hashlib
import gradio as gr
import pandas as pd
import fitz  # PyMuPDF for PDF processing
from PIL import Image
import pytesseract
from openpyxl import load_workbook

# Manually specify Tesseract path:
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

from sympy import sympify, simplify

# Optional: xlcalculator for Excel formula evaluation
try:
    from xlcalculator import ModelCompiler, Evaluator
    HAS_XLCALCULATOR = True
except ImportError:
    HAS_XLCALCULATOR = False

# LangChain + Community imports
from langchain_community.chat_models import ChatOpenAI
from langchain_community.embeddings import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.chains import RetrievalQA
from langchain.schema import HumanMessage

# Import your config (make sure it's in same folder)
from config import (
    OPENAI_API_KEY,
    VECTOR_STORE_PATH,
    UPLOAD_DIR,
    OCR_TEMP_DIR,
    ADMIN_USERNAME,
    ADMIN_PASSWORD,
    BASE_DIR
)

DEBUG = True
BYPASS_RETRIEVAL = False

# Global store for admin-uploaded content
admin_upload_content = ""

# Cache for PDF summaries to reduce repeated calls
pdf_summary_cache = {}

def debug_print(*args):
    if DEBUG:
        print(*args)

#####################################
# HELPER FUNCTIONS
#####################################

def read_file_data(file_obj):
    """
    Reads raw bytes from Gradio file object or string path.
    Returns (filename, raw_bytes).
    """
    if isinstance(file_obj, str):
        debug_print(f"[DEBUG] File object is a string path: {file_obj}")
        with open(file_obj, "rb") as f:
            data = f.read()
        name = os.path.basename(file_obj)
        return name, data

    if hasattr(file_obj, "read") and callable(file_obj.read):
        debug_print("[DEBUG] File object has .read()")
        data = file_obj.read()
        name = getattr(file_obj, "name", "uploaded_file")
        return name, data

    if hasattr(file_obj, "value"):
        debug_print("[DEBUG] File object has .value")
        data = file_obj.value
        name = getattr(file_obj, "name", "uploaded_file")
        return name, data

    if isinstance(file_obj, dict) and "name" in file_obj and "data" in file_obj:
        debug_print("[DEBUG] File object is dict with 'name' and 'data'")
        return file_obj["name"], file_obj["data"]

    raise ValueError(f"Unsupported file object type: {file_obj}")

def get_excel_evaluator(wb):
    try:
        from xlcalculator import ModelCompiler, Evaluator
        data = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_data = {}
            for row in ws.iter_rows():
                for cell in row:
                    sheet_data[cell.coordinate] = cell.value
            data[sheet] = sheet_data
        compiler = ModelCompiler()
        model = compiler.read_and_parse_dict(data)
        evaluator = Evaluator(model)
        return evaluator
    except Exception as e:
        debug_print("[DEBUG] Excel evaluator error:", e)
        return None

def advanced_understand_formula(formula_string):
    formula_clean = formula_string.lstrip('=').replace('^', '**')
    try:
        expr = sympify(formula_clean)
        simplified_expr = simplify(expr)
        return str(simplified_expr)
    except Exception as e:
        return f"Error in advanced understanding: {e}"

def summarize_text_chunk(text_chunk):
    prompt = (
        "Summarize the following engineering text concisely, "
        "focusing on key approvals, calculation methods, and important notes:\n\n"
        f"{text_chunk}\n\nSummary:"
    )
    messages = [HumanMessage(content=prompt)]
    try:
        llm = ChatOpenAI(temperature=0.1, openai_api_key=OPENAI_API_KEY, max_tokens=150)
        response = llm(messages)
        if response and hasattr(response, "content"):
            return response.content.strip()
        return ""
    except Exception as e:
        return f"Error summarizing chunk: {e}"

def advanced_understand_pdf(pdf_text):
    text_hash = hashlib.sha256(pdf_text.encode('utf-8')).hexdigest()
    if text_hash in pdf_summary_cache:
        debug_print("[DEBUG] Returning cached PDF summary.")
        return pdf_summary_cache[text_hash]

    max_chunk_size = 2000
    if len(pdf_text) <= max_chunk_size:
        summary = summarize_text_chunk(pdf_text)
    else:
        lines = pdf_text.splitlines()
        chunks = []
        current_chunk = ""
        for line in lines:
            if len(current_chunk) + len(line) + 1 < max_chunk_size:
                current_chunk += line + "\n"
            else:
                chunks.append(current_chunk)
                current_chunk = line + "\n"
        if current_chunk:
            chunks.append(current_chunk)
        debug_print(f"[DEBUG] Split PDF text into {len(chunks)} chunks for summarization.")

        chunk_summaries = []
        for chunk in chunks:
            summary_chunk = summarize_text_chunk(chunk)
            chunk_summaries.append(summary_chunk)

        combined_summary = "\n".join(chunk_summaries)
        summary = summarize_text_chunk(combined_summary)

    pdf_summary_cache[text_hash] = summary
    return summary

def extract_excel_content(file_path):
    debug_print(f"[DEBUG] Extracting Excel content from {file_path}")
    wb = load_workbook(file_path, data_only=False)
    evaluator = get_excel_evaluator(wb) if HAS_XLCALCULATOR else None

    content_list = []
    for sheet in wb.worksheets:
        sheet_content = f"Sheet: {sheet.title}\n"
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.data_type == 'f':  # cell formula
                    cell_text = f"Cell {cell.coordinate} Formula: {cell.value}"
                    adv = advanced_understand_formula(cell.value)
                    cell_text += f" | Advanced Understanding: {adv}"
                    if evaluator:
                        try:
                            result = evaluator.evaluate(f"{sheet.title}!{cell.coordinate}")
                            cell_text += f" | Evaluated Result: {result}"
                        except Exception as e:
                            cell_text += f" (Evaluation Error: {str(e)})"
                else:
                    cell_text = f"Cell {cell.coordinate} Value: {cell.value}"
                sheet_content += cell_text + "\n"
        content_list.append(sheet_content)

    return "\n".join(content_list)

def process_pdf_file(file_path):
    debug_print(f"[DEBUG] Processing PDF file: {file_path}")
    text = ""
    doc = fitz.open(file_path)
    for i in range(len(doc)):
        page = doc.load_page(i)
        page_text = page.get_text().strip()
        if page_text:
            text += page_text + "\n"
        else:
            pix = page.get_pixmap()
            img_path = os.path.join(OCR_TEMP_DIR, f"page_{i}.png")
            pix.save(img_path)
            img = Image.open(img_path)
            text += pytesseract.image_to_string(img) + "\n"
            os.remove(img_path)

    # Summarize with advanced approach
    advanced_summary = advanced_understand_pdf(text)
    combined_text = f"Extracted Text:\n{text}\n\nAdvanced PDF Understanding:\n{advanced_summary}"
    return combined_text

#####################################
# ADMIN MODE
#####################################

def generate_contextual_outcomes(admin_text, excel_content, pdf_content):
    """
    Creates a single "File Learning Outcomes" text, merging:
    - Admin Explanation
    - Summaries of Excel/PDF
    - Thought-provoking follow-up questions
    """
    combined = (
        f"Admin Explanation:\n{admin_text}\n\n"
        f"Excel Content:\n{excel_content}\n\n"
        f"PDF Content:\n{pdf_content}"
    )
    # Instead of generic suggestions, generate more context-specific prompts
    prompt = (
        "You are a specialized engineering AI. Below is new data (Excel calculations, PDF text, "
        "and the user's explanation). Provide a single, unified 'File Learning Outcome' that includes:\n"
        "1) A concise but detailed interpretation of the uploaded data.\n"
        "2) Specific, context-aware questions or clarifications if something seems ambiguous.\n"
        "3) Any alternative or more efficient methods to perform these calculations.\n\n"
        f"{combined}\n\n"
        "Please produce a direct, single text response labeled as 'File Learning Outcomes:' "
        "with your best interpretation and targeted follow-up prompts.\n"
    )

    messages = [HumanMessage(content=prompt)]

    def llm_call(msgs):
        llm = ChatOpenAI(temperature=0.2, openai_api_key=OPENAI_API_KEY)
        return llm(msgs)

    response_text = "Error: No response received."
    with concurrent.futures.ThreadPoolExecutor() as executor:
        future = executor.submit(llm_call, messages)
        try:
            response = future.result(timeout=20)
            if response and hasattr(response, "content"):
                response_text = response.content
        except Exception as e:
            debug_print("[DEBUG] Timeout/error generating contextual outcomes:", e)
            response_text = f"Error generating outcomes: {str(e)}"

    return response_text

def admin_learn(admin_text, excel_files, pdf_files):
    """
    Processes each file, merges into a single text, and returns a context-specific 'File Learning Outcomes'.
    We also return empty lists to clear the file components in the UI afterwards.
    """
    debug_print("[DEBUG] Entered admin_learn function.")
    excel_content = ""
    pdf_content = ""

    # Process Excel
    if excel_files:
        for fobj in excel_files:
            file_name, file_data = read_file_data(fobj)
            saved_path = os.path.join(UPLOAD_DIR, file_name)
            with open(saved_path, "wb") as fp:
                fp.write(file_data)
            excel_content_part = extract_excel_content(saved_path)
            excel_content += excel_content_part + "\n"
            os.remove(saved_path)

    # Process PDFs
    if pdf_files:
        for fobj in pdf_files:
            file_name, file_data = read_file_data(fobj)
            saved_path = os.path.join(UPLOAD_DIR, file_name)
            with open(saved_path, "wb") as fp:
                fp.write(file_data)
            pdf_content_part = process_pdf_file(saved_path)
            pdf_content += pdf_content_part + "\n"
            os.remove(saved_path)

    # Generate a single 'File Learning Outcomes' text
    outcomes = generate_contextual_outcomes(admin_text, excel_content, pdf_content)
    debug_print("[DEBUG] admin_learn -> outcome length:", len(outcomes))

    # Return outcomes + empty lists so file components auto-clear
    return outcomes, [], []

def confirm_learning(admin_text, excel_files, pdf_files):
    """
    Stores the new data into FAISS. Re-uses the same processing function for consistency.
    """
    debug_print("[DEBUG] Entered confirm_learning function.")
    outcomes, _, _ = admin_learn(admin_text, excel_files, pdf_files)
    global admin_upload_content
    admin_upload_content = outcomes

    embedding_model = OpenAIEmbeddings(openai_api_key=OPENAI_API_KEY)
    try:
        if os.path.exists(VECTOR_STORE_PATH):
            debug_print("[DEBUG] Loading existing FAISS store. (Dangerous deserialization).")
            vector_store = FAISS.load_local(
                VECTOR_STORE_PATH,
                embedding_model,
                allow_dangerous_deserialization=True
            )
            vector_store.add_texts([outcomes])
        else:
            debug_print("[DEBUG] Creating new FAISS store from scratch.")
            vector_store = FAISS.from_texts([outcomes], embedding_model)

        vector_store.save_local(VECTOR_STORE_PATH)
        debug_print("[DEBUG] Vector store updated successfully.")
    except Exception as e:
        debug_print("[DEBUG] Error updating vector store:", e)
        return f"Error updating vector store: {str(e)}", outcomes, [], []

    return "Learning confirmed and vector store updated!", outcomes, [], []

#####################################
# CLIENT MODE
#####################################

async def client_chat(query, history):
    """
    The user-facing chatbot. We do a retrieval-based approach if we have a vector store.
    The prompt is more context-aware to show deeper intelligence.
    """
    debug_print("[DEBUG] client_chat query:", query)
    start_time = time.time()
    final_answer = "No response"

    try:
        if (BYPASS_RETRIEVAL or not os.path.exists(VECTOR_STORE_PATH) or not admin_upload_content.strip()):
            debug_print("[DEBUG] No vector store => direct LLM call.")
            prompt = (
                f"You are a specialized engineering bot. The user asks:\n{query}\n\n"
                "Provide a detailed, step-by-step answer or calculation method. "
                "Reference known engineering standards if relevant."
            )
            messages = [HumanMessage(content=prompt)]

            def direct_llm_call(msgs):
                llm = ChatOpenAI(temperature=0.1, openai_api_key=OPENAI_API_KEY)
                return llm(msgs)

            try:
                response = await asyncio.to_thread(direct_llm_call, messages)
                final_answer = response.content if response else "No response"
            except Exception as e:
                final_answer = f"Error in direct call: {str(e)}"
        else:
            debug_print("[DEBUG] Using retrieval QA chain.")
            def run_retrieval_qa(q):
                embedding_model = OpenAIEmbeddings(openai_api_key=OPENAI_API_KEY)
                vector_store = FAISS.load_local(
                    VECTOR_STORE_PATH,
                    embedding_model,
                    allow_dangerous_deserialization=True
                )
                qa_chain = RetrievalQA.from_chain_type(
                    llm=ChatOpenAI(
                        temperature=0.1,
                        openai_api_key=OPENAI_API_KEY
                    ),
                    chain_type="stuff",
                    retriever=vector_store.as_retriever(),
                    verbose=True
                )

                # More context-aware prompt:
                # We can instruct the chain or rely on chain_type='stuff' that merges context + user query.
                return qa_chain.run(q)

            try:
                final_answer = await asyncio.to_thread(run_retrieval_qa, query)
            except Exception as e:
                final_answer = f"Error in retrieval QA: {str(e)}"

        history.append({"role": "user", "content": query})
        history.append({"role": "assistant", "content": final_answer})
    except Exception as e:
        debug_print("[DEBUG] Exception in client_chat:", e)
        history.append({"role": "system", "content": f"Error in client chat: {str(e)}"})

    elapsed = time.time() - start_time
    debug_print(f"[DEBUG] client_chat time: {elapsed:.2f}s")
    return history, history

#####################################
# GRADIO APP
#####################################

with gr.Blocks() as interface:
    gr.Markdown("# Advanced Engineering Chatbot\n")
    gr.Markdown("### Admin Mode vs. Client Mode with Enhanced Features\n")

    with gr.Tabs():
        with gr.Tab("Admin Mode (Private)"):
            gr.Markdown("**Upload & Learn** new engineering documents.")

            admin_text_input = gr.Textbox(
                label="Your Explanation/Context",
                lines=3,
                placeholder="Describe the methods, references, or context here..."
            )

            # Separate placeholders for each file type
            admin_excel_upload = gr.Files(
                label="Excel Files",
                file_types=[".xlsx", ".xls"]
            )
            admin_pdf_upload = gr.Files(
                label="PDF Files",
                file_types=[".pdf"]
            )

            process_btn = gr.Button("Process Uploads (Preview)")
            # Instead of separate combined preview + JSON suggestions, unify them
            file_learning_outcomes = gr.Textbox(
                label="File Learning Outcomes",
                lines=10
            )

            # Return typed file components so we can auto-clear them
            confirm_btn = gr.Button("Confirm and Learn")
            admin_learn_status = gr.Textbox(label="Learning Status", lines=2)

            # process_btn -> admin_learn
            # unify "File Learning Outcomes" output and auto-clear file components
            process_btn.click(
                fn=admin_learn,
                inputs=[admin_text_input, admin_excel_upload, admin_pdf_upload],
                outputs=[file_learning_outcomes, admin_excel_upload, admin_pdf_upload]
            )

            # confirm_btn -> confirm_learning
            confirm_btn.click(
                fn=confirm_learning,
                inputs=[admin_text_input, admin_excel_upload, admin_pdf_upload],
                outputs=[admin_learn_status, file_learning_outcomes, admin_excel_upload, admin_pdf_upload]
            )

        with gr.Tab("Client Mode"):
            gr.Markdown("**Query** the validated knowledge base here.")
            chatbot = gr.Chatbot(label="Engineering Chatbot", type="messages")
            query_input = gr.Textbox(label="Your Query")
            chat_state = gr.State([])

            send_btn = gr.Button("Send")
            send_btn.click(
                fn=client_chat,
                inputs=[query_input, chat_state],
                outputs=[chatbot, chat_state]
            )

# Optionally secure entire interface with admin credentials:
# interface.launch(auth=(ADMIN_USERNAME, ADMIN_PASSWORD), share=False)
interface.launch(share=False)
